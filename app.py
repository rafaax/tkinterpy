import sys
import pymysql
import datetime
import pprint
import csv
import time
import json
import os
import pytz
import chardet
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import requests
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from xhtml2pdf import pisa
from tkinter import *
from tkinter import messagebox, ttk, filedialog
from functions import Functions
from tkcalendar import Calendar, DateEntry
from src import conexao
from mysql.connector import connect


def submit():

    conn = connect(user=conexao.user, password=conexao.passw, host=conexao.host, database=conexao.db)
    
    data = input_date.get_date()
    placa = input_text.get()
    
    if not placa:
        messagebox.showwarning("Erro", "Por favor insira alguma placa válida.")
        return False

    placa = placa.replace('-', '') # removendo o traço da string para nao haver conflito

    data_hoje = now.strftime("%Y-%m-%d")
    data_input = data.strftime("%Y-%m-%d")
    
    if data_input > data_hoje:
        messagebox.showerror("Erro", "Data inserida não é valida por ser maior que a data atual!")
        return False

    cursor = conn.cursor()

    query = 'SELECT equi_id from sau_veiculos where placa = %s limit 1'
    cursor.execute(query, (placa, ))

    result = cursor.fetchone()

    if result:
        equipament_id = result[0]
        
        def close_loading():
            loading_page.destroy()
          
        loading_page = Toplevel()
        loading_page.title("Loading Page")
        loading_page.geometry("300x200")
        loading_page.resizable(False, False)

        progress_bar = ttk.Progressbar(loading_page, length=200, mode='indeterminate')
        progress_bar.pack(pady=50)

        progress_bar.start(20)

        query = (
            'SELECT placa, data_atualizacao, observacao, velocidade, pos_id, latitude, longitude FROM sau_posicionamento ' 
            'WHERE equi_id = %s AND date(data_atualizacao) = %s ORDER BY data_atualizacao DESC'
        )

        cursor.execute(query, (equipament_id, data_input))

        results = cursor.fetchall()

        

        if results: 
            loading_page.after(1000, close_loading)

            data_to_write = [result for result in results]
            
            filename_csv = 'files/csv/' + placa + str(equipament_id) + data_input + '.csv'

            with open(filename_csv, 'w', newline='', encoding='utf-8') as csvfile:
                csvwriter = csv.writer(csvfile)
                csvwriter.writerow(['PLACA', 'DATA', 'LOCAL', 'VELOCIDADE', 'VELOCIDADE VIA', 'LATITUDE', 'LONGITUDE', 'ULTRAPASSADO'])
                
                for row in data_to_write:
                    velocidade, pos_id = row[3], row[4]
                    row = list(row)
                    row.append(func.ultrapassado(velocidade, pos_id))
                    csvwriter.writerow(row)
            

            with open(filename_csv, 'rb') as f:
                charset = chardet.detect(f.read())
                
            encoding = charset['encoding']

            df = pd.read_csv(filename_csv, encoding=encoding)
            
            df['DATA'] = pd.to_datetime(df['DATA'])
            df['DATA'] = df['DATA'].dt.strftime('%d/%m/%Y %H:%M:%S')

            df['ULTRAPASSADO'] = df.apply(lambda row: func.ultrapassado(row['VELOCIDADE'], row['VELOCIDADE VIA']), axis=1)
            df[["VELOCIDADE", "VELOCIDADE VIA"]] = df[["VELOCIDADE", "VELOCIDADE VIA"]].map(func.add_km)

            font_style = {
                'font-family': 'Gotham Book',
                'font-size': '18px'
            }

            styled_df = df.style.set_properties(**font_style)

            styled_df = styled_df.map(func.velocidade_excedida, subset='ULTRAPASSADO')

            styled_df = styled_df.map(lambda x: f'color: {"black" if isinstance(x, str) else "purple"}')
    
            filename_xlsx = 'files/xlsx/' + placa + str(equipament_id) + data_input + '.xlsx'
            styled_df.to_excel(filename_xlsx, index=False, sheet_name = placa, engine='openpyxl')

            wb = load_workbook(filename_xlsx)
            ws = wb.active
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  
                ws.column_dimensions[column_letter].width = adjusted_width

            # wb.save(filename_xlsx)

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("XLSX", "*.xlsx")])
            if file_path:
                try:
                    wb.save(file_path)
                    if os.path.exists(file_path):
                        print("O arquivo foi salvo com sucesso.")
                        messagebox.showinfo("Sucesso", f"Arquivo inserido no caminho: {file_path} ")
                        return True
                    else:
                        print("Falha ao salvar o arquivo.")
                        messagebox.askokcancel("Erro", "ERRO AO SALVAR O ARQUIVO!!")
                        return True

                except Exception as e:
                    print(e)

        else:
            loading_page.after(1000, close_loading)
            messagebox.showerror("Erro", "Sem rota para este dia...")

        
    else:
        messagebox.showerror("Erro", "Placa não é valida!")
    




## main
## interface inicial

now = datetime.datetime.now()
func = Functions()

app = Tk()
app.title('Gerar Relatorio')
app.geometry('300x150')
app.wm_maxsize(width=600, height=400)
app.wm_minsize(width=300, height=150)

text_1 = Label(app, text="Insira a placa")
text_1.pack()

input_text = Entry(app, width=12)
input_text.pack()

input_date = DateEntry(app, width=12, background='black', foreground='white', borderwidth=2, year=now.year, date_pattern="dd/mm/yyyy")
input_date.pack(pady=10)

Button(app, text= "Salvar", command=submit).pack()


app.mainloop()